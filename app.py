from flask import Flask, Response, render_template, request, redirect, url_for, session, flash
import psycopg2
from werkzeug.security import check_password_hash
from functools import wraps
import re
import os
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from datetime import datetime
import getpass

# MAIN

dbuser = "postgres"
dbpwd = ""
dbport = "5432"


app = Flask(__name__)
# Ask user for dabase input
print("""Please input the following information.
        If values are not entered, the default will instead be used.""")
print("Database user (default postgres): ")
input_str = input().strip()
dbuser = input_str if len(input_str) > 0 else "postgres"
input_str = ''
while len(input_str) <= 0:
    # Source - https://stackoverflow.com/questions/52079846/how-to-hide-input-in-python-3-6
    # Posted by Muhammadabdulloh Komilov, modified by community. See post 'Timeline' for change history
    # Retrieved 2025-11-26, License - CC BY-SA 4.0
    input_str = getpass.getpass("Database password (required): ")
dbpwd = input_str
print("Port number (default 5432): ")
input_str = input().strip()
dbport = input_str if len(input_str) > 0 else "5432"
app.secret_key = "supersecretkey"
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
UPLOAD_FOLDER = os.path.join(os.getcwd(), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def get_db_connection():
    conn = psycopg2.connect(
        host="localhost",
        dbname="company_db",
        user=dbuser,
        password=dbpwd,
        port=dbport
    )
    return conn


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            flash("You do not have permission to perform this action.", "danger")
            return redirect(url_for("employees"))
        return f(*args, **kwargs)
    return decorated_function


def whitelist(input):
    # Source - https://stackoverflow.com/questions/56159087/how-can-i-whitelist-characters-from-a-string-in-python-3
    # Posted by Ajax1234
    # Retrieved 2025-11-25, License - CC BY-SA 4.0
    # Used as reference
    return re.sub('[^a-zA-Z0-9_ ]', '', input)


def is_valid(input):
    # Source - https://stackoverflow.com/questions/5698267/efficient-way-to-search-for-invalid-characters-in-python
    # Posted by ridgerunner, modified by community. See post 'Timeline' for change history
    # Retrieved 2025-11-25, License - CC BY-SA 3.0
    # Used as reference
    if re.search('[^a-zA-Z0-9_ ]', input):
        return False
    else:
        return True


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def validate_employee_row(row_data, row_num):
    """Validate a single employee row from Excel. Returns (is_valid, error_message)"""
    errors = []

    # Required fields
    required_fields = ['fname', 'minit', 'lname', 'ssn',
                       'address', 'sex', 'salary', 'dno', 'bdate', 'empdate']
    for field in required_fields:
        if field not in row_data or not row_data[field]:
            errors.append(f"Missing required field: {field}")

    if errors:
        return False, f"Row {row_num}: " + ", ".join(errors)

    # Validate SSN format (9 digits)
    if not re.match(r'^\d{9}$', str(row_data.get('ssn', '')).strip()):
        errors.append("SSN must be 9 digits")

    # Validate sex field
    if str(row_data.get('sex', '')).strip().upper() not in ['M', 'F']:
        errors.append("Sex must be 'M' or 'F'")

    # Validate salary is numeric
    try:
        int(row_data.get('salary', 0))
    except (ValueError, TypeError):
        errors.append("Salary must be a valid integer")

    # Validate dates
    for date_field in ['bdate', 'empdate']:
        date_val = row_data.get(date_field)
        if date_val:
            try:
                if isinstance(date_val, str):
                    datetime.strptime(date_val, '%Y-%m-%d')
            except ValueError:
                errors.append(f"{date_field} must be in YYYY-MM-DD format")

    if errors:
        return False, f"Row {row_num}: " + ", ".join(errors)

    return True, ""


def parse_excel_file(file_path, table_name):
    """Parse Excel file and return list of dictionaries for each row"""
    try:
        workbook = load_workbook(file_path)
        worksheet = workbook.active

        if not worksheet:
            return None, "No active worksheet found in Excel file"

        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())

        if not headers:
            return None, "No headers found in Excel file"

        # Parse data rows
        rows = []
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row_data = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row):
                    value = row[col_idx]
                    # Convert datetime objects to strings
                    if isinstance(value, datetime):
                        row_data[header] = value.strftime('%Y-%m-%d')
                    else:
                        row_data[header] = value

            if any(row_data.values()):  # Skip empty rows
                rows.append((row_idx, row_data))

        return rows, None
    except Exception as e:
        return None, f"Error reading Excel file: {str(e)}"


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT id, password_hash, role FROM app_user WHERE username = %s",
            (username,)
        )
        user = cur.fetchone()
        cur.close()
        conn.close()

        if user and check_password_hash(user[1], password):
            session["user_id"] = user[0]
            session["username"] = username
            session["role"] = user[2] if user[2] else "viewer"
            flash("Login successful!", "success")
            return redirect(url_for("employees"))
        else:
            flash("Invalid username or password", "danger")

    return render_template("login.html")


@app.route("/home")
@login_required
def home():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM employee;")
    total_employees = cur.fetchone()[0]
    cur.close()
    conn.close()

    return render_template(
        "home.html",
        username=session["username"],
        total_employees=total_employees
    )


def employee_query(search_name, selected_dept, sort_by, sort_dir):
    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT
            e.ssn,
            e.fname || ' ' || e.minit || '. ' || e.lname AS full_name,
            d.dname AS department,
            COUNT(DISTINCT dep.dependent_name) AS total_dependents,
            COUNT(DISTINCT p.pname) AS num_projects,
            STRING_AGG(DISTINCT p.pname, ', ') AS projects,
            COALESCE(SUM(w.hours), 0) AS total_hours
        FROM employee e
        LEFT JOIN department d ON e.dno = d.dnumber
        LEFT JOIN works_on w ON e.ssn = w.essn
        LEFT JOIN project p ON w.pno = p.pnumber
        LEFT JOIN dependent dep ON e.ssn = dep.essn
    """

    conditions = []
    params = []

    if search_name:
        conditions.append("(LOWER(e.fname) LIKE %s OR LOWER(e.lname) LIKE %s)")
        wl_search_name = whitelist(search_name)
        params.extend([f"%{wl_search_name.lower()}%",
                       f"%{wl_search_name.lower()}%"])

    if selected_dept:
        conditions.append("d.dname = %s")
        params.append(whitelist(selected_dept))

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += """
        GROUP BY e.ssn, e.fname, e.minit, e.lname, d.dname
    """

    if sort_by and sort_dir:
        sort_condition = " ORDER BY "
        sort_condition += "total_hours" if whitelist(
            sort_by) == "hours" else "e.fname"
        sort_condition += " DESC" if whitelist(sort_dir) == "desc" else " ASC"
        query += sort_condition
    query += ";"

    cur.execute(query, params)
    employees = cur.fetchall()

    cur.close()
    conn.close()

    return employees


@app.route("/employees/csv", methods=["GET"])
@login_required
def employees_csv():
    search_name = request.args.get("search_name", "").strip()
    selected_dept = request.args.get("department", "")
    sort_by = request.args.get("sort_by", "")
    sort_dir = request.args.get("sort_dir", "asc")

    employees = employee_query(search_name, selected_dept, sort_by, sort_dir)

    filename = whitelist(selected_dept).lower() + \
        "_employees" if selected_dept else "employees"
    emp_csv = "SSN,Name,Department,Dependents,Number of Projects,Projects,Total Hours\n"

    for emp in employees:
        str_emp = [str(value) for value in emp]
        str_emp[5] = str_emp[5].replace(',', '; ')
        row = ','.join(str_emp)
        emp_csv += row + "\n"

    return Response(
        emp_csv,
        mimetype="text/csv",
        headers={"Content-disposition": f"attachment; filename={filename}.csv"}
    )


@app.route("/employees", methods=["GET"])
@login_required
def employees():
    search_name = request.args.get("search_name", "").strip()
    selected_dept = request.args.get("department", "")
    sort_by = request.args.get("sort_by", "")
    sort_dir = request.args.get("sort_dir", "asc")

    employees = employee_query(search_name, selected_dept, sort_by, sort_dir)

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT DISTINCT dname FROM department ORDER BY dname;")
    departments = [row[0] for row in cur.fetchall()]

    cur.close()
    conn.close()

    return render_template(
        "employees.html",
        employees=employees,
        departments=departments,
        selected_dept=selected_dept,
        search_name=search_name,
        sort_by=sort_by,
        sort_dir=sort_dir
    )


@app.route("/projects", methods=["GET"])
@login_required
def projects():

    conn = get_db_connection()
    cur = conn.cursor()

    search_name = request.args.get("search_name", "").strip()
    selected_dept = request.args.get("department", "")
    sort_by = request.args.get("sort_by", "")
    sort_dir = request.args.get("sort_dir", "asc")

    query = """
        SELECT
            p.pnumber,
            p.pname,
            d.dname AS department,
            COUNT(DISTINCT w.essn) AS total_employees,
            COALESCE(SUM(w.hours), 0) AS total_hours
        FROM project p
        LEFT JOIN department d ON p.dnum = d.dnumber
        LEFT JOIN works_on w ON p.pnumber = w.pno
    """

    conditions = []
    params = []

    if search_name:
        conditions.append("LOWER(p.pname) LIKE %s")
        params.append(f"%{search_name.lower()}%")

    if selected_dept:
        conditions.append("d.dname = %s")
        params.append(selected_dept)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " GROUP BY p.pnumber, p.pname, d.dname "

    if sort_by and sort_dir:
        sort_condition = " ORDER BY "
        sort_condition += "total_hours" if whitelist(
            sort_by) == "hours" else "total_employees"
        sort_condition += " DESC" if whitelist(sort_dir) == "desc" else " ASC"
        query += sort_condition
    query += ";"

    cur.execute(query, params)
    projects = cur.fetchall()

    cur.execute("SELECT DISTINCT dname FROM department ORDER BY dname;")
    departments = [row[0] for row in cur.fetchall()]

    cur.close()
    conn.close()

    return render_template(
        "projects.html",
        projects=projects,
        departments=departments,
        search_name=search_name,
        selected_dept=selected_dept,
        sort_by=sort_by,
        sort_dir=sort_dir
    )


@app.route("/projects/<pid>", methods=["GET", "POST"])
@login_required
def project(pid):

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        if session.get("role") != "admin":
            flash("You do not have permission to modify project assignments.",
                  "danger")
            return redirect(url_for("project", pid=pid))
        try:
            emp_id = request.form["emp_id"]
            hours = request.form["hours"]
            cur.execute(
                """
                INSERT INTO Works_On (Essn, Pno, Hours)
                VALUES (%s, %s, %s)
                ON CONFLICT (Essn, Pno)
                DO UPDATE SET Hours = Works_On.Hours + EXCLUDED.Hours;
                """,
                (emp_id, pid, hours)
            )
            conn.commit()
            flash("Hours updated for employee on this project.", "success")
        except Exception:
            conn.rollback()
            flash(
                "Error: Ensure you selected an employee and hours are between 0-999.",
                "danger"
            )

    cur.execute(
        """
        SELECT Fname, Minit, Lname, Hours
        FROM Works_on
        INNER JOIN Employee ON Employee.Ssn = Works_on.Essn
        WHERE Pno = %s
        """,
        (pid,)
    )
    emps_on_project = cur.fetchall()

    cur.execute("SELECT Ssn, Fname, Minit, Lname FROM Employee ORDER BY Fname")
    employees = cur.fetchall()

    cur.execute("SELECT Pname FROM Project WHERE Pnumber = %s", (pid,))
    proj_name = cur.fetchone()

    cur.close()
    conn.close()
    return render_template(
        "project.html",
        proj_name=proj_name,
        emps_on_project=emps_on_project,
        employees=employees
    )


@app.route("/managers")
@login_required
def managers():

    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT
            d.dname,
            d.dnumber,
            COALESCE(e.fname || ' ' || e.minit || '. ' || e.lname, 'None') AS manager_name,
            COUNT(DISTINCT emp.ssn) AS employee_count,
            COALESCE(SUM(w.hours), 0) AS total_hours
        FROM department d
        LEFT JOIN employee e ON d.mgr_ssn = e.ssn
        LEFT JOIN employee emp ON emp.dno = d.dnumber
        LEFT JOIN works_on w ON w.essn = emp.ssn
        GROUP BY d.dname, d.dnumber, manager_name
        ORDER BY d.dnumber;
    """

    cur.execute(query)
    rows = cur.fetchall()

    cur.close()
    conn.close()

    return render_template("managers.html", managers=rows)


@app.route("/employee/add", methods=["GET", "POST"])
@admin_required
def add_employee():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT dnumber, dname FROM department ORDER BY dname;")
    departments = cur.fetchall()

    cur.execute(
        "SELECT ssn, fname || ' ' || lname FROM employee ORDER BY fname;")
    supervisors = cur.fetchall()

    if request.method == "POST":
        fname = request.form["fname"]
        minit = request.form["minit"]
        lname = request.form["lname"]
        ssn = request.form["ssn"]
        address = request.form["address"]
        sex = request.form["sex"]
        salary = request.form["salary"]
        super_ssn = request.form["super_ssn"] or None
        dno = request.form["dno"]
        bdate = request.form["bdate"]
        empdate = request.form["empdate"]

        try:
            cur.execute("""
                INSERT INTO employee
                (fname, minit, lname, ssn, address, sex, salary,
                 super_ssn, dno, bdate, empdate)
                VALUES (%s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s)
            """, (fname, minit, lname, ssn, address, sex, salary,
                  super_ssn, dno, bdate, empdate))

            conn.commit()
            flash("Employee added successfully!", "success")
            return redirect(url_for("employees"))

        except psycopg2.errors.UniqueViolation:
            conn.rollback()
            flash(
                "Error: An employee with the provided SSN already exists",
                "danger")
        except Exception as e:
            conn.rollback()
            flash(f"Error adding employee: {e}", "danger")

    cur.close()
    conn.close()

    return render_template("add_employee.html",
                           departments=departments,
                           supervisors=supervisors)


@app.route("/employee/<ssn>/edit", methods=["GET", "POST"])
@admin_required
def edit_employee(ssn):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM employee WHERE ssn = %s;", (ssn,))
    emp = cur.fetchone()

    if not emp:
        flash("Employee not found.", "danger")
        cur.close()
        conn.close()
        return redirect(url_for("employees"))

    cur.execute("SELECT dnumber, dname FROM department ORDER BY dname;")
    departments = cur.fetchall()

    cur.execute(
        "SELECT ssn, fname || ' ' || minit || '. ' || lname "
        "FROM employee WHERE ssn <> %s ORDER BY fname;",
        (ssn,)
    )
    supervisors = cur.fetchall()

    if request.method == "POST":
        address = request.form["address"]
        salary = request.form["salary"]
        dno = request.form["dno"]
        super_ssn = request.form["super_ssn"] or None

        try:
            cur.execute("""
                UPDATE employee
                SET address = %s, salary = %s, dno = %s, super_ssn = %s
                WHERE ssn = %s;
            """, (address, salary, dno, super_ssn, ssn))

            conn.commit()
            flash("Employee updated successfully!", "success")
            return redirect(url_for("employees"))

        except Exception as e:
            conn.rollback()
            flash(f"Error updating employee: {e}", "danger")

    cur.close()
    conn.close()

    return render_template("edit_employee.html",
                           emp=emp,
                           departments=departments,
                           supervisors=supervisors)


@app.route("/employee/<ssn>/delete")
@admin_required
def delete_employee(ssn):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT 1 FROM department WHERE mgr_ssn = %s LIMIT 1;", (ssn,))
    dept_mgr = cur.fetchone()

    if dept_mgr:
        flash("Cannot delete employee: They are a department manager.",
              "danger")
        cur.close()
        conn.close()
        return redirect(url_for("employees"))

    cur.execute("SELECT 1 FROM employee WHERE super_ssn = %s LIMIT 1;",
                (ssn,))
    emp_mgr = cur.fetchone()

    if emp_mgr:
        flash("Cannot delete employee: They are a supervisor.", "danger")
        cur.close()
        conn.close()
        return redirect(url_for("employees"))

    try:
        cur.execute("DELETE FROM employee WHERE ssn = %s;", (ssn,))
        conn.commit()
        flash("Employee deleted.", "info")
    except psycopg2.DatabaseError as e:
        conn.rollback()
        detail = e.pgerror
        if "works_on" in e.pgerror:
            detail = "Employee is assigned to a project"
        if "dependent" in e.pgerror:
            detail = "Employee has dependents"
        flash(f"Could not delete employee: {detail}.", "danger")
    except Exception as e:
        conn.rollback()
        flash("Error deleting employee: " + str(e), "danger")

    cur.close()
    conn.close()
    return redirect(url_for("employees"))


# @app.route("/import", methods=["GET", "POST"])
# @admin_required
# def import_data():
#     if request.method == "POST":

#         if 'file' not in request.files:
#             flash("No file selected", "danger")
#             return redirect(url_for("import_data"))

#         file = request.files['file']
#         table_name = request.form.get('table', '').strip().lower()

#         if file.filename == '':
#             flash("No file selected", "danger")
#             return redirect(url_for("import_data"))

#         if not allowed_file(file.filename):
#             flash("Invalid file format. Only .xlsx files are allowed", "danger")
#             return redirect(url_for("import_data"))

#         if table_name not in ['employee', 'project', 'department', 'dependent', 'works_on']:
#             flash("Invalid table selected", "danger")
#             return redirect(url_for("import_data"))

#         filename = secure_filename(file.filename)
#         file_path = os.path.join(UPLOAD_FOLDER, filename)
#         file.save(file_path)

#         try:

#             rows, parse_error = parse_excel_file(file_path, table_name)
#             if parse_error:
#                 flash(parse_error, "danger")
#                 return redirect(url_for("import_data"))

#             if not rows:
#                 flash("No data found in Excel file", "warning")
#                 return redirect(url_for("import_data"))

#             conn = get_db_connection()
#             cur = conn.cursor()

#             successful_rows = 0
#             failed_rows = []

#             for row_num, row_data in rows:
#                 try:
#                     if table_name == 'employee':
#                         is_valid, error_msg = validate_employee_row(
#                             row_data, row_num)
#                         if not is_valid:
#                             failed_rows.append(error_msg)
#                             continue

#                         cur.execute(
#                             "SELECT 1 FROM employee WHERE ssn = %s", (row_data['ssn'],))
#                         if cur.fetchone():
#                             failed_rows.append(
#                                 f"Row {row_num}: Employee with SSN {row_data['ssn']} already exists")
#                             continue

#                         cur.execute(
#                             """
#                             INSERT INTO employee
#                             (fname, minit, lname, ssn, address, sex, salary, super_ssn, dno, bdate, empdate)
#                             VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
#                             """,
#                             (
#                                 row_data.get('fname'),
#                                 row_data.get('minit', ''),
#                                 row_data.get('lname'),
#                                 str(row_data.get('ssn')).strip(),
#                                 row_data.get('address'),
#                                 str(row_data.get('sex', '')).upper(),
#                                 int(row_data.get('salary', 0)),
#                                 None,
#                                 int(row_data.get('dno', 0)),
#                                 row_data.get('bdate'),
#                                 row_data.get('empdate')
#                             )
#                         )

#                         successful_rows += 1

#                     elif table_name == 'project':
#                         if not all(row_data.get(field) for field in ['pname', 'pnumber', 'plocation', 'dnum']):
#                             failed_rows.append(
#                                 f"Row {row_num}: Missing required fields")
#                             continue

#                         cur.execute(
#                             "SELECT 1 FROM project WHERE pnumber = %s", (int(row_data['pnumber']),))
#                         if cur.fetchone():
#                             failed_rows.append(
#                                 f"Row {row_num}: Project with number {row_data['pnumber']} already exists")
#                             continue

#                         cur.execute("""
#                             INSERT INTO project (pname, pnumber, plocation, dnum)
#                             VALUES (%s, %s, %s, %s)
#                         """, (
#                             row_data.get('pname'),
#                             int(row_data.get('pnumber')),
#                             row_data.get('plocation'),
#                             int(row_data.get('dnum'))
#                         ))
#                         successful_rows += 1

#                     elif table_name == 'department':

#                         if not all(row_data.get(field) for field in ['dname', 'dnumber']):
#                             failed_rows.append(
#                                 f"Row {row_num}: Missing required fields (dname, dnumber)")
#                             continue

#                         cur.execute(
#                             "SELECT 1 FROM department WHERE dnumber = %s", (int(row_data['dnumber']),))
#                         if cur.fetchone():
#                             failed_rows.append(
#                                 f"Row {row_num}: Department with number {row_data['dnumber']} already exists")
#                             continue

#                         cur.execute("""
#                             INSERT INTO department (dname, dnumber, mgr_ssn)
#                             VALUES (%s, %s, %s)
#                         """, (
#                             row_data.get('dname'),
#                             int(row_data.get('dnumber')),
#                             row_data.get('mgr_ssn')
#                         ))
#                         successful_rows += 1

#                     elif table_name == 'dependent':
#                         if not all(row_data.get(field) for field in ['essn', 'dependent_name', 'sex', 'bdate', 'relationship']):
#                             failed_rows.append(
#                                 f"Row {row_num}: Missing required fields")
#                             continue

#                         cur.execute("""
#                             INSERT INTO dependent (essn, dependent_name, sex, bdate, relationship)
#                             VALUES (%s, %s, %s, %s, %s)
#                         """, (
#                             row_data.get('essn'),
#                             row_data.get('dependent_name'),
#                             str(row_data.get('sex', '')).upper(),
#                             row_data.get('bdate'),
#                             row_data.get('relationship')
#                         ))
#                         successful_rows += 1

#                     elif table_name == 'works_on':
#                         if not all(row_data.get(field) for field in ['essn', 'pno', 'hours']):
#                             failed_rows.append(
#                                 f"Row {row_num}: Missing required fields (essn, pno, hours)")
#                             continue

#                         cur.execute("""
#                             INSERT INTO works_on (essn, pno, hours)
#                             VALUES (%s, %s, %s)
#                             ON CONFLICT (essn, pno) DO UPDATE SET hours = works_on.hours + EXCLUDED.hours
#                         """, (
#                             row_data.get('essn'),
#                             int(row_data.get('pno')),
#                             float(row_data.get('hours', 0))
#                         ))
#                         successful_rows += 1

#                 except psycopg2.errors.ForeignKeyViolation as e:
#                     conn.rollback()
#                     failed_rows.append(
#                         f"Row {row_num}: Foreign key constraint violated - {str(e)}")
#                 except psycopg2.errors.UniqueViolation as e:
#                     conn.rollback()
#                     failed_rows.append(
#                         f"Row {row_num}: Duplicate entry - record already exists")
#                 except Exception as e:
#                     conn.rollback()
#                     failed_rows.append(f"Row {row_num}: {str(e)}")

#             conn.commit()
#             cur.close()
#             conn.close()

#             if successful_rows > 0:
#                 flash(
#                     f"Successfully imported {successful_rows} rows into {table_name} table", "success")

#             if failed_rows:
#                 flash(f"Failed to import {len(failed_rows)} rows. Errors: " + " | ".join(failed_rows[:5]) +
#                       (f" and {len(failed_rows) - 5} more..." if len(failed_rows) > 5 else ""), "warning")

#             return redirect(url_for("import_data"))

#         except Exception as e:
#             flash(f"Error processing Excel file: {str(e)}", "danger")
#             return redirect(url_for("import_data"))
#         finally:
#             if os.path.exists(file_path):
#                 os.remove(file_path)

#     return render_template("import_data.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


if __name__ == "__main__":
    app.run(debug=True)
